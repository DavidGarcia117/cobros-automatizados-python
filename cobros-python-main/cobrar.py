import os
import openpyxl
from util import Mail as m
from util import Whatsapp as w
import pyautogui as pg

# Cargar archivo Excel
archivo ='C:/Users/RSOC_TIC/Documents/cobros por correo/cobros por correo/cobros-python-main/datos_envio_cobro_SENSIBLE.xlsx'
wb = openpyxl.load_workbook(archivo)
sheet = wb.active  # Obtén la hoja activa (primera hoja por defecto)

# Parámetros de correo
correoCopia = "david.garcia@supergiroscasanare.co"
correoOrigen = "david.garcia@supergiroscasanare.co"
contra = "Casio2025*"

# Función para validar campos
def validar_datos(name, deuda, txtValor, direccion, numWhatsapp, correoDestino):
    if not all([name, deuda, txtValor, direccion, numWhatsapp, correoDestino]):
        return False
    if not isinstance(numWhatsapp, str):  # Validación del número como cadena
        return False
    return True

# Bucle para iterar sobre las filas del archivo
for i in range(2, sheet.max_row + 1):  # Comenzar desde la segunda fila
    try:
        # Leer valores de la fila actual
        name = sheet.cell(row=i, column=1).value
        deuda = sheet.cell(row=i, column=2).value
        txtValor = sheet.cell(row=i, column=3).value
        direccion = sheet.cell(row=i, column=4).value
        numWhatsapp = str(sheet.cell(row=i, column=5).value) if sheet.cell(row=i, column=5).value else None
        correoDestino = sheet.cell(row=i, column=6).value

        # Detener si el campo clave (name) está vacío
        if not name:  # O verifica otra columna clave
            print(f"Fila {i}: Campo 'name' vacío. Finalizando el proceso.")
            break

        if validar_datos(name, deuda, txtValor, direccion, numWhatsapp, correoDestino):
            # Enviar mensajes
            w.enviarWhatsapp(numWhatsapp, name, correoDestino, direccion, txtValor, deuda)
            m.enviarCorreo(correoDestino, correoCopia, correoOrigen, contra, numWhatsapp, name, direccion, txtValor, deuda)
            print(f"Mensajes enviados correctamente a {name}")
        else:
            print(f"Datos incompletos para {name} en fila {i}.")
            pg.alert(f"Faltan datos en el usuario {name}")
    except Exception as e:
        print(f"Error procesando fila {i}: {str(e)}")
